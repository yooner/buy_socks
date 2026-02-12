"""
运行所有股票的阈值策略回测
生成 Excel 报告，对比收益率，满足条件时自动打 git tag
"""

import os
import sys
import json
import pandas as pd
import subprocess
import io
from datetime import datetime
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import FormulaRule

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from buy_socks_sell_thresholds import (
    INITIAL_CAPITAL,
    buy_levels,
    buy_ratios,
    sell_thresholds,
    sell_ratios,
    run_backtest
)

CACHE_DIR = "cache"
EXCEL_FILE = "stocks_results.xlsx"


def get_cached_stocks() -> List[str]:
    """获取缓存目录中所有股票代码"""
    stocks = []
    for filename in os.listdir(CACHE_DIR):
        if filename.endswith("_daily.json"):
            stock_code = filename.replace("_daily.json", "")
            stocks.append(stock_code)
    return sorted(stocks)


def load_existing_results() -> pd.DataFrame:
    """从 Excel 加载现有结果"""
    if not os.path.exists(EXCEL_FILE):
        return pd.DataFrame()

    try:
        return pd.read_excel(EXCEL_FILE)
    except Exception as e:
        print(f"加载 Excel 失败: {e}")
        return pd.DataFrame()


def run_backtest_for_stock(stock_code: str) -> Tuple[Optional[float], Dict[int, float]]:
    """对单个股票运行回测，捕获输出并提取收益率"""
    print(f"\n{'='*80}")
    print(f"正在回测 {stock_code}...")
    print(f"{'='*80}")

    old_stdout = sys.stdout
    captured_output = io.StringIO()

    try:
        sys.stdout = captured_output
        total_return, yearly_returns = run_backtest(stock_code)
        sys.stdout = old_stdout

        output = captured_output.getvalue()

        if total_return is None:
            print(f"  ❌ {stock_code}: 数据不足，跳过")
            return None, {}

        print(f"  ✅ {stock_code}: 总收益率 {total_return:+.2f}%")
        if yearly_returns:
            year_str = ', '.join([f"{y}:{v:+.1f}%" for y, v in sorted(yearly_returns.items())])
            print(f"     年度收益: {year_str}")

        return total_return, yearly_returns

    except Exception as e:
        sys.stdout = old_stdout
        print(f"  ❌ {stock_code}: 回测失败 - {e}")
        import traceback
        traceback.print_exc()
        return None, {}

    finally:
        sys.stdout = old_stdout


def save_results_to_excel(results: List[Dict], previous_results: pd.DataFrame = None):
    """保存结果到 Excel"""
    if os.path.exists(EXCEL_FILE):
        try:
            os.remove(EXCEL_FILE)
        except Exception as e:
            print(f"删除旧文件失败: {e}")

    columns = ['编号', '总收益率']

    all_years = set()
    for r in results:
        if r['yearly_returns']:
            all_years.update(r['yearly_returns'].keys())

    sorted_years = sorted(all_years)
    for year in sorted_years:
        columns.append(f'{year}年收益率')

    columns.append('gittag')

    df = pd.DataFrame(columns=columns)

    for r in results:
        row = {
            '编号': r['stock_code'],
            '总收益率': r['total_return'] / 100
        }

        for year in sorted_years:
            row[f'{year}年收益率'] = r['yearly_returns'].get(year, 0) / 100

        row['gittag'] = r.get('gittag', '')

        df = pd.concat([df, pd.DataFrame([row])], ignore_index=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "回测结果"

    for r_idx, row_data in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            col_name = df.columns[c_idx - 1]
            if '收益率' in col_name:
                cell.number_format = '0.00%'
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )

    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws.row_dimensions[1].height = 25

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 14

    for year_idx, year in enumerate(sorted_years):
        col_letter = chr(ord('C') + year_idx)
        ws.column_dimensions[col_letter].width = 14

    last_col = chr(ord('C') + len(sorted_years))
    ws.column_dimensions[last_col].width = 20

    for row_idx in range(2, len(results) + 2):
        ws.row_dimensions[row_idx].height = 20

    wb.save(EXCEL_FILE)
    print(f"\n结果已保存到 {EXCEL_FILE}")

    return df


def check_and_create_git_tag(results: List[Dict], previous_df: pd.DataFrame):
    """检查是否需要创建 git tag"""
    if previous_df is None or previous_df.empty:
        print("\n没有历史数据，跳过 git tag 检查")
        return

    improved_count = 0
    total_count = 0

    for r in results:
        stock_code = r['stock_code']
        current_return = r['total_return']

        if pd.isna(current_return):
            continue

        prev_row = previous_df[previous_df['编号'] == stock_code]
        if prev_row.empty:
            continue

        prev_return = prev_row['总收益率'].values[0]
        if pd.isna(prev_return):
            continue

        total_count += 1
        if current_return > prev_return:
            improved_count += 1

    if total_count == 0:
        print("\n没有可比对的股票，跳过 git tag 检查")
        return

    improvement_rate = improved_count / total_count * 100
    print(f"\n收益率改善股票数: {improved_count}/{total_count} ({improvement_rate:.1f}%)")

    if improvement_rate >= 60:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        tag_name = f"v{timestamp}"
        commit_msg = datetime.now().strftime("%H:%M:%S")

        try:
            subprocess.run(['git', 'add', EXCEL_FILE], check=True)
            subprocess.run(['git', 'commit', '-m', commit_msg], check=True)
            subprocess.run(['git', 'tag', tag_name], check=True)
            print(f"\n✅ 已创建 git tag: {tag_name}")
            print(f"   Commit message: {commit_msg}")
        except subprocess.CalledProcessError as e:
            print(f"\n❌ Git 操作失败: {e}")
    else:
        print(f"\n收益率改善比例 ({improvement_rate:.1f}%) 未达到 60%，不创建 git tag")


def main():
    stocks = get_cached_stocks()
    print(f"找到 {len(stocks)} 只缓存股票: {', '.join(stocks)}")
    print(f"\n策略参数:")
    print(f"  买入档位: {buy_levels}")
    print(f"  买入比例: {buy_ratios}")
    print(f"  卖出档位: {sell_thresholds}")
    print(f"  卖出比例: {sell_ratios}")
    print(f"  起始资金: {INITIAL_CAPITAL}")

    previous_df = load_existing_results()
    if not previous_df.empty:
        print(f"\n已加载历史数据，共 {len(previous_df)} 条记录")

    results = []

    for i, stock_code in enumerate(stocks, 1):
        total_return, yearly_returns = run_backtest_for_stock(stock_code)

        if total_return is None:
            continue

        result = {
            'stock_code': stock_code,
            'total_return': total_return,
            'yearly_returns': yearly_returns
        }
        results.append(result)

    if not results:
        print("\n没有有效的回测结果")
        return

    df = save_results_to_excel(results, previous_df)

    check_and_create_git_tag(results, previous_df)

    print(f"\n{'='*60}")
    print("回测完成!")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
