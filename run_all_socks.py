"""
通用股票回测脚本
支持运行多种策略，按策略名称生成不同的结果文件
"""

import os
import sys
import json
import pandas as pd
import subprocess
import io
import time
import argparse
from datetime import datetime
from typing import Dict, List, Optional, Tuple, Any

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# 从ana_stocks导入tushare配置
try:
    from ana_stocks import TS_TOKEN, pro as tushare_pro
except ImportError:
    TS_TOKEN = None
    tushare_pro = None

STRATEGIES = {
    'thresholds': {
        'module': 'buy_socks_sell_thresholds',
        'params': ['INITIAL_CAPITAL', 'buy_levels', 'buy_ratios', 'sell_atr_multipliers', 'sell_ratios'],
        'run_backtest': 'run_backtest',
        'name': '阈值策略',
        'excel_suffix': 'thresholds'
    },
    'outbreak': {
        'module': 'buy_socks_sell_outbreak',
        'params': ['INITIAL_CAPITAL'],
        'run_backtest': 'run_backtest',
        'name': '趋势爆发策略',
        'excel_suffix': 'outbreak'
    },
    'trend': {
        'module': 'buy_socks_sell_trend',
        'params': ['INITIAL_CAPITAL'],
        'run_backtest': 'run_backtest',
        'name': '趋势结构策略',
        'excel_suffix': 'trend'
    },
    'bodong': {
        'module': 'buy_socks_sell_bodong',
        'params': ['INITIAL_CAPITAL'],
        'run_backtest': 'run_backtest',
        'name': '波动率策略',
        'excel_suffix': 'bodong'
    }
}

CACHE_DIR = "cache"


def get_cached_stocks() -> List[str]:
    """获取缓存目录中所有股票代码"""
    stocks = []
    for filename in os.listdir(CACHE_DIR):
        if filename.endswith("_daily.json"):
            stock_code = filename.replace("_daily.json", "")
            stocks.append(stock_code)
    return sorted(stocks)


def get_sh_index_price() -> Optional[float]:
    """获取上证指数最新价格（优先从缓存文件，否则从tushare获取并缓存）"""
    sh_index_file = os.path.join(CACHE_DIR, "sh000001_daily.json")
    
    # 首先尝试从缓存文件读取
    if os.path.exists(sh_index_file):
        try:
            with open(sh_index_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            if data and len(data) > 0:
                # 获取最新一天的数据
                latest_data = data[-1]
                close_price = latest_data.get('close', latest_data.get('收盘', None))
                if close_price is not None:
                    return float(close_price)
        except Exception as e:
            print(f"⚠️  读取上证指数缓存失败: {e}")
    
    # 缓存不存在或读取失败，从tushare获取
    print("📡 从tushare获取上证指数数据...")
    try:
        # 使用从ana_stocks导入的token和pro
        if TS_TOKEN and tushare_pro:
            pro = tushare_pro
            print(f"✅ 使用ana_stocks中的tushare配置")
        else:
            # 备选：尝试从环境变量获取
            import tushare as ts
            token = os.getenv('TUSHARE_TOKEN')
            if not token:
                print("⚠️  未找到tushare token，请确保ana_stocks.py中有TS_TOKEN配置")
                return None
            ts.set_token(token)
            pro = ts.pro_api()
        
        # 获取上证指数日线数据（使用index_daily接口，最近100天）
        df = pro.index_daily(ts_code='000001.SH', limit=100)
        
        if df is None or len(df) == 0:
            print("⚠️  从tushare获取上证指数数据失败")
            return None
        
        # 转换数据格式并缓存
        df = df.sort_values('trade_date')  # 按日期升序排列
        data_list = []
        for _, row in df.iterrows():
            data_list.append({
                'date': row['trade_date'],
                'open': float(row['open']),
                'high': float(row['high']),
                'low': float(row['low']),
                'close': float(row['close']),
                'volume': float(row['vol']),
                'amount': float(row['amount'])
            })
        
        # 保存到缓存文件
        with open(sh_index_file, 'w', encoding='utf-8') as f:
            json.dump(data_list, f, ensure_ascii=False, indent=2)
        
        print(f"✅ 上证指数数据已缓存到: {sh_index_file}")
        
        # 返回最新收盘价
        return float(df.iloc[-1]['close'])
        
    except ImportError:
        print("⚠️  未安装tushare，请运行: pip install tushare")
        return None
    except Exception as e:
        print(f"⚠️  从tushare获取上证指数失败: {e}")
        return None


def get_position_buy_count_by_sh_index(sh_index_price: Optional[float]) -> int:
    """根据上证指数点位决定分仓数量
    
    规则：
    - 3000以下：分4仓
    - 3000~3300：分3仓
    - 3300~3600：分2仓
    - 3600+：不分仓（全仓买入）
    
    Returns:
        分仓数量，0表示不分仓（全仓买入）
    """
    if sh_index_price is None:
        print("⚠️  无法获取上证指数，使用默认分仓数量4")
        return 4
    
    if sh_index_price < 3000:
        count = 4
    elif sh_index_price < 3300:
        count = 3
    elif sh_index_price < 3600:
        count = 2
    else:
        count = 0  # 3600+ 不分仓
    
    print(f"📊 上证指数: {sh_index_price:.2f}，分仓数量: {count if count > 0 else '不分仓(全仓)'}")
    return count


def get_position_buy_count_by_year(year: int) -> int:
    """根据年份获取分仓数量（基于历史上证指数区间）
    
    历史区间划分：
    - 2016: 3000以下 -> 4仓
    - 2017: 3000~3300 -> 3仓
    - 2018~2019: 3000以下 -> 4仓
    - 2020: 3000~3300 -> 3仓
    - 2021: 3300~3600 -> 2仓
    - 2022: 3000~3300 -> 3仓
    - 2023: 3000~3300 -> 3仓
    - 2024: 3300~3600 -> 2仓
    - 2025: 3300~3600 -> 2仓 (上半年), 3600+ -> 0仓 (下半年)
    
    Returns:
        分仓数量，0表示不分仓（全仓买入）
    """
    # 年份到分仓数量的映射
    year_config = {
        2016: 3,  # 3000以下
        2017: 2,  # 3000~3300
        2018: 3,  # 3000以下
        2019: 3,  # 3000以下
        2020: 2,  # 3000~3300
        2021: 0,  # 3300~3600
        2022: 2,  # 3000~3300
        2023: 2,  # 3000~3300
        2024: 0,  # 3300~3600
        2025: 0,  # 3600+
        2026: 0,
    }
    
    count = year_config.get(year, 4)  # 默认4仓
    
    # 打印年份和对应的上证指数区间
    level_desc = {
        4: "3000以下",
        3: "3000~3300",
        2: "3300~3600",
        0: "3600+"
    }
    print(f"📅 年份: {year}，上证指数区间: {level_desc.get(count, '未知')}，分仓数量: {count if count > 0 else '不分仓(全仓)'}")
    
    return count


def load_existing_results(excel_file: str) -> pd.DataFrame:
    """从 Excel 加载现有结果"""
    if not os.path.exists(excel_file):
        return pd.DataFrame()

    try:
        df = pd.read_excel(excel_file, dtype={'编号': str})
        return df
    except Exception as e:
        print(f"加载 Excel 失败: {e}")
        return pd.DataFrame()


def run_backtest_for_stock(
    strategy_key: str,
    stock_code: str,
    outbreak_sell_e_buyback_threshold: float = None,
    outbreak_position_buy_count: int = None
) -> Tuple[Optional[float], Dict[int, float], Dict[str, Any]]:
    """对单个股票运行指定策略的回测
    
    Args:
        strategy_key: 策略标识
        stock_code: 股票代码
        outbreak_sell_e_buyback_threshold: 卖出E后的买回阈值（仅bodong策略有效）
        outbreak_position_buy_count: 爆发买入分仓数量（仅bodong策略有效）
    """
    strategy = STRATEGIES[strategy_key]
    module_name = strategy['module']

    print(f"\n{'='*80}")
    print(f"策略: {strategy['name']} | 股票: {stock_code}")
    print(f"{'='*80}")

    old_stdout = sys.stdout
    captured_output = io.StringIO()

    try:
        module = __import__(module_name, fromlist=[strategy['run_backtest']])
        run_backtest_func = getattr(module, strategy['run_backtest'])

        sys.stdout = captured_output
        # 对于bodong策略，传入额外参数
        if strategy_key == 'bodong':
            result = run_backtest_func(
                stock_code,
                outbreak_sell_e_buyback_threshold=outbreak_sell_e_buyback_threshold,
                outbreak_position_buy_count=outbreak_position_buy_count
            )
        else:
            result = run_backtest_func(stock_code)
        sys.stdout = old_stdout

        if result is None:
            print(f"  ❌ {stock_code}: 回测失败，返回None")
            return None, {}, {}

        if isinstance(result, tuple) and len(result) >= 2:
            total_return, yearly_returns = result[0], result[1]
            extra_data = result[2] if len(result) > 2 else {}
        else:
            total_return, yearly_returns = result, {}
            extra_data = {}

        if total_return is None:
            print(f"  ❌ {stock_code}: 数据不足，跳过")
            return None, {}, {}

        print(f"  ✅ {stock_code}: 总收益率 {total_return:+.2f}%")
        if yearly_returns:
            year_str = ', '.join([f"{y}:{v:+.1f}%" for y, v in sorted(yearly_returns.items())])
            print(f"     年度收益: {year_str}")

        return total_return, yearly_returns, extra_data

    except Exception as e:
        sys.stdout = old_stdout
        print(f"  ❌ {stock_code}: 回测失败 - {e}")
        import traceback
        traceback.print_exc()
        return None, {}, {}

    finally:
        sys.stdout = old_stdout


def save_results_to_excel(results: List[Dict], excel_file: str, strategy_key: str, previous_df: pd.DataFrame = None, enable_update: bool = True):
    """保存结果到 Excel
    enable_update: 是否允许更新Excel（仅当60%以上股票收益率改善时才更新）
    """
    if not enable_update:
        if previous_df is not None and not previous_df.empty:
            print("\n收益率改善不足60%，不更新Excel，保留历史最优数据")
        else:
            print(f"\n首次运行，保存新数据到 {excel_file}")

        if os.path.exists(excel_file):
            print(f"   现有文件: {excel_file}")
        return previous_df if previous_df is not None else pd.DataFrame()

    columns = ['编号', '总收益率']

    all_years = set()
    for r in results:
        if r.get('yearly_returns'):
            all_years.update(r['yearly_returns'].keys())

    sorted_years = sorted(all_years)
    for year in sorted_years:
        columns.append(f'{year}年收益率')

    columns.append('gittag')

    # 收集所有行数据，然后一次性创建DataFrame（避免pd.concat警告）
    rows = []
    for r in results:
        row = {
            '编号': str(r['stock_code']),
            '总收益率': r['total_return'] / 100 if r['total_return'] is not None else 0
        }

        for year in sorted_years:
            row[f'{year}年收益率'] = r.get('yearly_returns', {}).get(year, 0) / 100 if r.get('yearly_returns') else 0

        row['gittag'] = r.get('gittag', '')
        rows.append(row)

    df = pd.DataFrame(rows, columns=columns)

    df['编号'] = df['编号'].astype(str)

    if os.path.exists(excel_file):
        try:
            os.remove(excel_file)
        except Exception as e:
            print(f"删除旧文件失败: {e}")

    wb = Workbook()
    ws = wb.active
    ws.title = "回测结果"

    for r_idx, row_data in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            col_name = df.columns[c_idx - 1]
            if '收益率' in col_name:
                cell.number_format = '0.00%'
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )

    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws.row_dimensions[1].height = 25

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 14

    for year_idx, year in enumerate(sorted_years):
        col_letter = chr(ord('C') + year_idx)
        ws.column_dimensions[col_letter].width = 14

    last_col = chr(ord('C') + len(sorted_years))
    ws.column_dimensions[last_col].width = 20

    for row_idx in range(2, len(results) + 2):
        ws.row_dimensions[row_idx].height = 20

    wb.save(excel_file)
    print(f"\n结果已保存到 {excel_file}")

    return df


def check_and_create_git_tag(results: List[Dict], previous_df: pd.DataFrame, strategy_key: str):
    """检查是否需要创建 git tag，并显示详细的收益率对比
    返回: (improvement_rate, enable_update) - 改善比例和是否允许更新Excel
    """
    if previous_df is None or previous_df.empty:
        print("\n没有历史数据，首次运行")
        return 100.0, True

    print("\n" + "="*60)
    print("收益率对比详情")
    print("="*60)
    print(f"{'股票代码':<12} {'上次收益率':>14} {'本次收益率':>14} {'变化':>14}")
    print("-"*60)

    improved_count = 0
    declined_count = 0
    unchanged_count = 0
    total_count = 0

    for r in results:
        stock_code = str(r['stock_code'])
        current_return = r['total_return']

        if current_return is None or pd.isna(current_return):
            continue

        prev_row = previous_df[previous_df['编号'].astype(str) == stock_code]
        if prev_row.empty:
            print(f"{stock_code:<12} {'无历史数据':>14} {current_return:>13.2f}% {'--':>14}")
            continue

        prev_return = prev_row['总收益率'].values[0]
        if pd.isna(prev_return):
            print(f"{stock_code:<12} {'无历史数据':>14} {current_return:>13.2f}% {'--':>14}")
            continue

        total_count += 1
        change = current_return - prev_return

        if change > 0.001:
            improved_count += 1
            change_str = f"+{change:.2f}%"
        elif change < -0.001:
            declined_count += 1
            change_str = f"{change:.2f}%"
        else:
            unchanged_count += 1
            change_str = "0.00%"

        print(f"{stock_code:<12} {prev_return:>13.2f}% {current_return:>13.2f}% {change_str:>14}")

    print("-"*60)
    print(f"总计: {total_count} 只股票 | 改善: {improved_count} | 下降: {declined_count} | 持平: {unchanged_count}")
    print("="*60)

    if total_count == 0:
        print("\n没有可比对的股票，首次运行")
        return 100.0, True

    improvement_rate = improved_count / total_count * 100
    print(f"\n收益率改善股票数: {improved_count}/{total_count} ({improvement_rate:.1f}%)")

    if improvement_rate >= 60:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        tag_name = f"{STRATEGIES[strategy_key]['excel_suffix']}_{timestamp}"
        commit_msg = datetime.now().strftime("%H:%M:%S")

        try:
            subprocess.run(['git', 'add', '-A'], check=True, capture_output=True)
            subprocess.run(['git', 'commit', '-m', commit_msg], check=True, capture_output=True)
            subprocess.run(['git', 'tag', tag_name], check=True, capture_output=True)
            print(f"\n✅ 已创建 git tag: {tag_name}")
            print(f"   Commit message: {commit_msg}")
        except subprocess.CalledProcessError as e:
            print(f"\n❌ Git 操作失败: {e}")

        return improvement_rate, True
    else:
        print(f"\n收益率改善比例 ({improvement_rate:.1f}%) 未达到 60%，不创建 git tag")
        return improvement_rate, False


def run_strategy(
    strategy_key: str,
    stocks: List[str],
    enable_git_tag: bool = True,
    outbreak_sell_e_buyback_threshold: float = None,
    outbreak_position_buy_count: int = None
):
    """运行单个策略
    
    Args:
        strategy_key: 策略标识
        stocks: 股票代码列表
        enable_git_tag: 是否启用git tag
        outbreak_sell_e_buyback_threshold: 卖出E后的买回阈值（仅bodong策略有效）
        outbreak_position_buy_count: 爆发买入分仓数量（仅bodong策略有效）
    """
    strategy = STRATEGIES[strategy_key]
    excel_suffix = strategy['excel_suffix']
    excel_file = f"socks_results_{excel_suffix}.xlsx"

    module = __import__(strategy['module'], fromlist=[''])
    params = {}
    for param_name in strategy['params']:
        if hasattr(module, param_name):
            params[param_name] = getattr(module, param_name)

    print(f"\n{'='*60}")
    print(f"策略: {strategy['name']}")
    print(f"{'='*60}")
    print(f"策略参数:")
    for k, v in params.items():
        print(f"  {k}: {v}")
    # 打印bodong策略的额外参数
    if strategy_key == 'bodong':
        if outbreak_sell_e_buyback_threshold is not None:
            print(f"  outbreak_sell_e_buyback_threshold: {outbreak_sell_e_buyback_threshold}")
        if outbreak_position_buy_count is not None:
            print(f"  outbreak_position_buy_count: {outbreak_position_buy_count}")
    print(f"{'='*60}")

    previous_df = load_existing_results(excel_file)
    if not previous_df.empty:
        print(f"\n已加载历史数据，共 {len(previous_df)} 条记录")

    results = []
    holding_stocks = []  # 收集持仓股票信息
    all_downtrend_to_rally_trades = []  # 收集所有股票的下降趋势→自然回升买入记录

    for i, stock_code in enumerate(stocks, 1):
        total_return, yearly_returns, extra_data = run_backtest_for_stock(
            strategy_key,
            stock_code,
            outbreak_sell_e_buyback_threshold=outbreak_sell_e_buyback_threshold,
            outbreak_position_buy_count=outbreak_position_buy_count
        )

        if total_return is None:
            continue
        
        # 每只股票回测后暂停0.5秒，降低CPU负载
        # time.sleep(0.3)

        result = {
            'stock_code': stock_code,
            'total_return': total_return,
            'yearly_returns': yearly_returns
        }
        results.append(result)
        
        # 收集持仓信息
        if extra_data and 'holding_info' in extra_data and extra_data['holding_info']:
            holding_info = extra_data['holding_info']
            holding_stocks.append({
                'stock_code': stock_code,
                'position': holding_info['position'],
                'cost_price': holding_info['cost_price'],
                'current_price': holding_info['current_price'],
                'price_change_pct': holding_info['price_change_pct'],
                'start_date': holding_info.get('start_date', 'N/A'),
                'holding_days': holding_info.get('holding_days', 0)
            })
        
        # 收集下降趋势→自然回升买入记录
        if extra_data and 'downtrend_to_rally_trades' in extra_data and extra_data['downtrend_to_rally_trades']:
            for trade in extra_data['downtrend_to_rally_trades']:
                trade['stock_code'] = stock_code
                all_downtrend_to_rally_trades.append(trade)

    if not results:
        print(f"\n没有有效的回测结果")
        return
    
    # 打印持仓股票信息（仅对波动策略）
    if strategy_key == 'bodong' and holding_stocks:
        print(f"\n{'='*110}")
        print("持仓股票信息")
        print(f"{'='*110}")
        print(f"{'股票代码':<12} {'持仓数量':>10} {'持仓价格':>12} {'最新价格':>12} {'涨跌幅':>10} {'开始日期':>12} {'持仓天数':>10}")
        print("-"*110)
        for h in holding_stocks:
            start_date_str = str(h['start_date'])[:10] if h['start_date'] else 'N/A'
            print(f"{h['stock_code']:<12} {h['position']:>10} {h['cost_price']:>12.2f} {h['current_price']:>12.2f} {h['price_change_pct']:>+9.2f}% {start_date_str:>12} {h['holding_days']:>10}天")
        print(f"{'='*110}")
        print(f"总计: {len(holding_stocks)} 只股票处于持仓状态")
    
    # 打印所有股票的下降趋势→自然回升买入统计（仅对波动策略）
    if strategy_key == 'bodong' and all_downtrend_to_rally_trades:
        print(f"\n{'='*110}")
        print("【所有股票下降趋势→自然回升买入统计】")
        print(f"{'='*110}")
        print(f"总买入次数: {len(all_downtrend_to_rally_trades)}")
        total_profit = sum(t['profit'] for t in all_downtrend_to_rally_trades)
        avg_profit_pct = sum(t['profit_pct'] for t in all_downtrend_to_rally_trades) / len(all_downtrend_to_rally_trades)
        print(f"总盈亏: {total_profit:+.2f} 元")
        print(f"平均收益率: {avg_profit_pct:+.2f}%")
        print(f"\n详细记录:")
        print(f"{'序号':<6} {'股票代码':<10} {'买入日期':<12} {'买入价格':>10} {'卖出日期':<12} {'卖出价格':>10} {'盈亏':>12} {'收益率':>10}")
        print("-"*110)
        for i, t in enumerate(all_downtrend_to_rally_trades, 1):
            print(f"{i:<6} {t['stock_code']:<10} {t['buy_date']:<12} {t['buy_price']:>10.2f} {t['sell_date']:<12} {t['sell_price']:>10.2f} {t['profit']:>+12.2f} {t['profit_pct']:>+9.2f}%")
        print(f"{'='*110}")

    if enable_git_tag:
        _, enable_update = check_and_create_git_tag(results, previous_df, strategy_key)
    else:
        enable_update = True

    save_results_to_excel(results, excel_file, strategy_key, previous_df, enable_update)

    return results


def main():
    parser = argparse.ArgumentParser(description='通用股票回测脚本')
    parser.add_argument('--strategies', type=str, default='all',
                        help='要运行的策略，用逗号分隔，如: thresholds,outbreak 或 all')
    parser.add_argument('--stocks', type=str, default=None,
                        help='要回测的股票代码，用逗号分隔，如: 000002,603496')
    parser.add_argument('--all', action='store_true',
                        help='运行所有缓存股票（将启用 git tag 功能）')
    # bodong策略专用参数
    parser.add_argument('--outbreak-sell-e-buyback-threshold', type=float, default=None,
                        help='卖出E后的买回阈值（仅bodong策略有效）')
    parser.add_argument('--outbreak-position-buy-count', type=int, default=None,
                        help='爆发买入分仓数量（仅bodong策略有效，默认根据上证指数自动计算：3000以下4仓，3000-3300分3仓，3300-3600分2仓，3600+不分仓）')
    args = parser.parse_args()

    all_mode = args.all
    enable_git_tag = all_mode

    if all_mode:
        stocks = get_cached_stocks()
        print(f"✅ ALL 模式：将运行所有 {len(stocks)} 只缓存股票")
        print(f"   Git Tag 功能：{'启用' if enable_git_tag else '禁用'}")
    elif args.stocks:
        stocks = [s.strip() for s in args.stocks.split(',') if s.strip()]
        stocks = [s for s in stocks if len(s) >= 6]
        print(f"📋 自定义模式：运行 {len(stocks)} 只指定股票")
        print(f"   Git Tag 功能：禁用")
    else:
        stocks = get_cached_stocks()
        print(f"📋 默认模式：运行所有 {len(stocks)} 只缓存股票")
        print(f"   Git Tag 功能：禁用（使用 --all 启用）")

    print(f"\n缓存股票: {', '.join(stocks)}")

    if args.strategies == 'all':
        strategies_to_run = list(STRATEGIES.keys())
    else:
        strategies_to_run = [s.strip() for s in args.strategies.split(',') if s.strip() in STRATEGIES]

    if not strategies_to_run:
        print(f"\n可用策略: {', '.join(STRATEGIES.keys())}")
        return

    print(f"\n将运行策略: {', '.join([STRATEGIES[s]['name'] for s in strategies_to_run])}")

    # 获取上证指数点位，根据点位决定分仓数量（如果命令行没有指定）
    sh_index_price = get_sh_index_price()
    auto_position_buy_count = get_position_buy_count_by_sh_index(sh_index_price)
    
    # 如果命令行指定了分仓数量，使用命令行的；否则使用根据上证指数自动计算的
    if args.outbreak_position_buy_count is not None:
        final_position_buy_count = args.outbreak_position_buy_count
        print(f"⚙️  使用命令行指定的分仓数量: {final_position_buy_count}")
    else:
        final_position_buy_count = auto_position_buy_count if auto_position_buy_count > 0 else None
        if final_position_buy_count is None:
            print(f"⚙️  上证指数{sh_index_price:.2f} >= 3600，不分仓（全仓买入）")
        else:
            print(f"⚙️  使用根据上证指数自动计算的分仓数量: {final_position_buy_count}")

    all_results = {}
    for strategy_key in strategies_to_run:
        results = run_strategy(
            strategy_key,
            stocks,
            enable_git_tag=enable_git_tag,
            outbreak_sell_e_buyback_threshold=args.outbreak_sell_e_buyback_threshold,
            outbreak_position_buy_count=final_position_buy_count
        )
        if results:
            all_results[strategy_key] = results

    print(f"\n{'='*60}")
    print("所有回测完成!")
    print(f"{'='*60}")

    if all_results:
        print(f"\n策略结果汇总:")
        for strategy_key, results in all_results.items():
            avg_return = sum(r['total_return'] for r in results) / len(results)
            win_count = sum(1 for r in results if r['total_return'] > 0)
            print(f"  {STRATEGIES[strategy_key]['name']}: 平均收益率 {avg_return:+.2f}% ({win_count}/{len(results)} 盈利)")

    if enable_git_tag:
        print(f"\n✅ Git Tag 已按策略分别创建")


if __name__ == "__main__":
    main()
